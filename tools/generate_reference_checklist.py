"""
Generate an Excel checklist of which reference images exist in assets/reference/.

For each unit × rank × appearance-state combination the cell shows:
  ✓   — image exists
  (blank) — missing

Run:
    .venv\Scripts\python.exe tools\generate_reference_checklist.py
Output:
    data/reference_checklist.xlsx
"""

import sqlite3
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

import openpyxl
from openpyxl.styles import (Alignment, Font, PatternFill, Border, Side)
from openpyxl.utils import get_column_letter

from src.recognition.template_matcher import _parse_reference_filename

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

_REF_DIR  = _ROOT / "assets" / "reference"
_DB_PATH  = _ROOT / "data" / "sheet_exports" / "unit_meta.db"
_OUT_PATH = _ROOT / "data" / "sheet_exports" / "reference_checklist.xlsx"

_RANKS = list(range(1, 8))   # 1–7

# States that apply to every unit (always shown as columns)
_UNIVERSAL_STATES = ["base", "max"]

# States that only apply to specific units
_UNIT_SPECIFIC_STATES: dict[str, list[str]] = {
    "scrapper": ["absorbed", "scrapemall", "absorbed_scrapemall"],
    "trapper":  ["huntmaster"],
}

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------

_GREEN  = "FF92D050"   # has image
_RED    = "FFFFC7CE"   # missing
_HEADER = "FF4472C4"   # dark blue header
_SUBHDR = "FFB4C6E7"   # light blue sub-header
_ALT    = "FFF2F2F2"   # alternate row tint

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


# ---------------------------------------------------------------------------
# Scan reference library
# ---------------------------------------------------------------------------

def scan_library(ref_dir: Path) -> dict[tuple[str, str, int], str]:
    """
    Returns a dict keyed by (unit_id, appearance_state, rank) → filename.
    Skips hero_portraits and talent_icons sub-dirs.
    """
    found: dict[tuple[str, str, int], str] = {}
    skip = {"hero_portraits", "talent_icons"}
    for unit_dir in sorted(ref_dir.iterdir()):
        if not unit_dir.is_dir() or unit_dir.name in skip:
            continue
        unit_id = unit_dir.name
        for img in sorted(unit_dir.glob("*.png")):
            entry = _parse_reference_filename(unit_id, img)
            if entry is None:
                continue
            key = (unit_id, entry.appearance_state, entry.merge_rank)
            found[key] = img.name
    return found


# ---------------------------------------------------------------------------
# Load units from DB
# ---------------------------------------------------------------------------

def load_units(db_path: Path) -> list[tuple[str, str]]:
    """Returns [(unit_id, display_name), ...] sorted by display_name."""
    conn = sqlite3.connect(db_path)
    rows = conn.execute(
        "SELECT unit_id, display_name FROM units ORDER BY display_name"
    ).fetchall()
    conn.close()
    return [(r[0], r[1]) for r in rows]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def build_workbook(units: list[tuple[str, str]],
                   library: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reference Checklist"

    # ------------------------------------------------------------------
    # Build column layout
    # Columns: Unit Name | unit_id | base r1..r7 | max r1..r7 |
    #          [per-unit special state r1..r7 …]
    # We collect all unique special states across all units first.
    # ------------------------------------------------------------------

    all_special_states: list[str] = []
    for states in _UNIT_SPECIFIC_STATES.values():
        for s in states:
            if s not in all_special_states:
                all_special_states.append(s)

    # Column groups: (state_label, [rank, rank, ...])
    col_groups: list[tuple[str, list[int]]] = []
    for state in _UNIVERSAL_STATES + all_special_states:
        col_groups.append((state, _RANKS))

    # Flatten columns
    # Fixed columns first: Unit Name, unit_id
    fixed_cols = ["Unit Name", "unit_id"]
    rank_cols: list[tuple[str, int]] = []   # (state, rank)
    for state, ranks in col_groups:
        for r in ranks:
            rank_cols.append((state, r))

    total_cols = len(fixed_cols) + len(rank_cols)

    # ------------------------------------------------------------------
    # Row 1 — state group headers (merged across rank columns)
    # ------------------------------------------------------------------
    ws.cell(1, 1, "Unit Name").font = Font(bold=True, color="FFFFFFFF")
    ws.cell(1, 1).fill = _fill(_HEADER)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 1).border = _THIN

    ws.cell(1, 2, "unit_id").font = Font(bold=True, color="FFFFFFFF")
    ws.cell(1, 2).fill = _fill(_HEADER)
    ws.cell(1, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 2).border = _THIN

    col = 3
    for state, ranks in col_groups:
        start_col = col
        end_col   = col + len(ranks) - 1
        label = state.replace("_", " ").title()
        ws.cell(1, start_col, label)
        ws.cell(1, start_col).font      = Font(bold=True, color="FFFFFFFF")
        ws.cell(1, start_col).fill      = _fill(_HEADER)
        ws.cell(1, start_col).alignment = Alignment(horizontal="center",
                                                     vertical="center")
        ws.cell(1, start_col).border = _THIN
        if end_col > start_col:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1,   end_column=end_col,
            )
        col = end_col + 1

    ws.row_dimensions[1].height = 20

    # ------------------------------------------------------------------
    # Row 2 — rank sub-headers
    # ------------------------------------------------------------------
    ws.cell(2, 1, "").fill = _fill(_SUBHDR)
    ws.cell(2, 2, "").fill = _fill(_SUBHDR)

    col = 3
    for state, ranks in col_groups:
        for r in ranks:
            c = ws.cell(2, col, f"R{r}")
            c.font      = Font(bold=True)
            c.fill      = _fill(_SUBHDR)
            c.alignment = Alignment(horizontal="center")
            c.border    = _THIN
            col += 1

    ws.row_dimensions[2].height = 16

    # Freeze rows 1–2 and the first two columns
    ws.freeze_panes = "C3"

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    for row_idx, (unit_id, display_name) in enumerate(units, start=3):
        fill = _fill(_ALT) if row_idx % 2 == 0 else PatternFill()

        # Unit name
        c = ws.cell(row_idx, 1, display_name)
        c.font      = Font(bold=True)
        c.fill      = fill
        c.alignment = Alignment(vertical="center")
        c.border    = _THIN

        # unit_id
        c = ws.cell(row_idx, 2, unit_id)
        c.fill      = fill
        c.alignment = Alignment(vertical="center")
        c.font      = Font(color="FF808080", italic=True)
        c.border    = _THIN

        col = 3
        for state, ranks in col_groups:
            # Determine if this state is applicable to this unit
            applicable = (
                state in _UNIVERSAL_STATES
                or unit_id in _UNIT_SPECIFIC_STATES
                and state in _UNIT_SPECIFIC_STATES[unit_id]
            )

            for r in ranks:
                key = (unit_id, state, r)
                has = key in library

                if not applicable:
                    # Grey out non-applicable cells
                    c = ws.cell(row_idx, col, "N/A")
                    c.fill      = _fill("FFD9D9D9")
                    c.font      = Font(color="FF808080", italic=True)
                    c.alignment = Alignment(horizontal="center",
                                            vertical="center")
                elif has:
                    c = ws.cell(row_idx, col, "✓")
                    c.fill      = _fill(_GREEN)
                    c.font      = Font(bold=True, color="FF375623")
                    c.alignment = Alignment(horizontal="center",
                                            vertical="center")
                    c.comment   = None
                else:
                    c = ws.cell(row_idx, col, "")
                    c.fill      = _fill(_RED)
                    c.alignment = Alignment(horizontal="center",
                                            vertical="center")

                c.border = _THIN
                col += 1

        ws.row_dimensions[row_idx].height = 18

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------
    ws.column_dimensions["A"].width = 22   # Unit Name
    ws.column_dimensions["B"].width = 20   # unit_id

    col = 3
    for state, ranks in col_groups:
        for _ in ranks:
            ws.column_dimensions[get_column_letter(col)].width = 5
            col += 1

    # ------------------------------------------------------------------
    # Summary row at top (after data)
    # ------------------------------------------------------------------
    summary_row = len(units) + 3
    ws.cell(summary_row, 1, "TOTAL COVERED").font = Font(bold=True)
    ws.cell(summary_row, 1).border = _THIN

    col = 3
    for state, ranks in col_groups:
        for r in ranks:
            count = sum(
                1 for uid, _ in units
                if (uid, state, r) in library
            )
            c = ws.cell(summary_row, col, count)
            c.font      = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
            c.border    = _THIN
            col += 1

    # ------------------------------------------------------------------
    # Legend sheet
    # ------------------------------------------------------------------
    leg = wb.create_sheet("Legend")
    leg["A1"] = "Colour"
    leg["B1"] = "Meaning"
    leg["A1"].font = Font(bold=True)
    leg["B1"].font = Font(bold=True)

    leg["A2"].fill = _fill(_GREEN)
    leg["B2"] = "Reference image exists in assets/reference/"

    leg["A3"].fill = _fill(_RED)
    leg["B3"] = "Missing — need to label and promote from data/to_label/"

    leg["A4"].fill = _fill("FFD9D9D9")
    leg["B4"] = "N/A — state does not apply to this unit"

    leg.column_dimensions["A"].width = 12
    leg.column_dimensions["B"].width = 50

    return wb


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    print("Scanning reference library...")
    library = scan_library(_REF_DIR)
    print(f"  Found {len(library)} reference images.")

    print("Loading units from database...")
    units = load_units(_DB_PATH)
    print(f"  Found {len(units)} units.")

    print("Building workbook...")
    wb = build_workbook(units, library)

    _OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(_OUT_PATH)
    print(f"\nSaved: {_OUT_PATH}")
    print("Open the file in Excel to see what's covered and what's missing.")


if __name__ == "__main__":
    main()